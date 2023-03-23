import openpyxl


class ReadExcel:

    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def read_cases(self):
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        res = list(sh.rows)
        title = [i.value for i in res[0]]
        # print(title)
        cases = []
        for case in res[1:]:
            data = [i.value for i in case]
            cases.append(dict(zip(title, data)))
        return cases

    def write_data(self, row, column, value):
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        sh.cell(row=row, column=column, value=value)
        workbook.save(self.filename)


if __name__ == '__main__':
    cases = ReadExcel(r"C:\Users\siyaowu\Desktop\pythonproject\ningmengban\py35_14day\testcase\testcase.xlsx",
                      'register')
    res = cases.read_cases()
    print(res)
